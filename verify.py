"""
QC 검증 API 라우터
파일 업로드 → 검증 실행(백그라운드 스레드) → 결과 조회 → Excel 다운로드
"""
import io
import json
import os
import urllib.parse
import secrets
import threading
import time
import logging
from pathlib import Path

from fastapi import APIRouter, Header, HTTPException, UploadFile
from fastapi.responses import (FileResponse, JSONResponse, Response,
                               StreamingResponse)

from pydantic import BaseModel

from backend import config
from backend.routers.auth import require_login
from backend.services import verifier as svc
from backend.services import excel_report, feedback_store, result_store, rule_catalog

logger = logging.getLogger(__name__)
router = APIRouter()

# ─── 인메모리 상태 (단일 사용자) ─────────────────────────────────────────────
_lock = threading.Lock()
_state: dict = {
    "status":       "idle",   # idle | running | done | error
    "progress":     0,
    "message":      "대기 중",
    "error":        None,
    "result":       None,
    "excel_status": "idle",   # idle | generating | ready | error
    "excel_error":  None,
}
# ─── 큰 파일 내려주기 ─────────────────────────────────────────────────────────
#
# `FileResponse`를 쓰지 않는다. 신버전 Starlette는 `Range` 요청에 206으로
# 답하는데, **크기를 먼저 재고(`stat`) 나중에 파일을 연다.** 그 사이에 결과
# Excel이 다시 만들어져 `os.replace`로 갈아 끼워지면 선언한 Content-Length보다
# 적게 보내게 되고, h11이 이렇게 죽는다:
#
#     GET /api/result/excel?t=… 206 Partial Content
#     ERROR: Exception in ASGI application
#     h11._util.LocalProtocolError: Too little data for declared Content-Length
#
# 브라우저에는 「다운로드할 수 없음」으로만 보인다.
#
# 그래서 **먼저 열고, 그 핸들에서 크기를 잰다.** 실제로 보낼 바이트 수와
# Content-Length가 같은 출처에서 나오므로 어긋날 수가 없다. 파일이 도중에
# 갈아 끼워져도 열어 둔 핸들은 원래 내용을 계속 준다.
#
# `Accept-Ranges: none`으로 **조각내기 자체를 막는다** — 브라우저가 여러 연결로
# 나눠 받으면 그만큼 어긋날 창이 넓어진다. 사내망이라 이어받기는 포기해도 된다.
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def stream_xlsx(path: Path, filename: str) -> StreamingResponse:
    fh = open(path, "rb")
    size = os.fstat(fh.fileno()).st_size

    def chunks():
        try:
            while True:
                b = fh.read(1024 * 1024)
                if not b:
                    break
                yield b
        finally:
            fh.close()

    # 한글 파일명은 ASCII 헤더에 그대로 못 담는다 — RFC 5987로 함께 싣는다.
    quoted = urllib.parse.quote(filename)
    return StreamingResponse(chunks(), media_type=_XLSX, headers={
        "Content-Length": str(size),
        "Accept-Ranges": "none",
        "Content-Disposition": f"attachment; filename*=UTF-8''{quoted}",
        # no-store가 없으면 브라우저가 응답을 캐시해, 로그아웃한 뒤에도
        # 네트워크 요청 없이 이전 파일을 그대로 내준다.
        "Cache-Control": "no-store, no-cache, must-revalidate",
    })


# ─── 다운로드 전용 «일회용 표» ────────────────────────────────────────────────
#
# 결과 Excel은 `<a download>`로 받는다 — `fetch().blob()`은 파일 크기만큼
# 브라우저 메모리를 잡고, 서버가 이미 무거운 상태면 그게 마지막 한 방이 된다.
# 브라우저가 디스크로 바로 흘려보내게 하는 편이 크기와 무관하게 안전하다.
#
# 그런데 `<a download>`는 **커스텀 헤더를 실을 수 없다.** 사번을 URL에 담으면
# 서버 접근 로그·브라우저 방문 기록·Referer에 그대로 남는데, 이 시스템은
# 사번이 유일한 인증 수단이라 그게 곧 자격증명이다.
#
# 그래서 사번 대신 **30초짜리 일회용 표**를 싣는다. 로그에 남더라도 그때는
# 이미 만료됐거나 소진된 뒤다.
_TICKETS: dict[str, tuple[str, float]] = {}
_TICKET_TTL = 30.0


def _issue_ticket(emp_id: str) -> str:
    now = time.time()
    with _lock:
        for k, (_, exp) in list(_TICKETS.items()):   # 만료분 청소
            if exp < now:
                _TICKETS.pop(k, None)
        tok = secrets.token_urlsafe(24)
        _TICKETS[tok] = (emp_id, now + _TICKET_TTL)
    return tok


def _check_ticket(tok: str) -> str | None:
    """살아 있으면 사번, 아니면 None. **소진하지 않는다.**

    처음에는 «한 번 쓰면 없어지는» 표로 만들었는데 브라우저와 맞지 않았다.
    Edge는 큰 파일을 **여러 연결로 나눠** 받는다 — 실측:

        GET ?t=…  range=-              -> 200   첫 요청이 표를 먹는다
        GET ?t=…  range=bytes=52428800- -> 401
        GET ?t=…  range=bytes=104857600- -> 401

    그래서 진행률이 끝까지 갔다가 「다운로드할 수 없음 — 권한 부여 필요」로
    끝난다. 파일은 이미 받았는데 이어지는 조각이 거절당한 것이다.

    그래서 **시간만 제한한다.** 창(30초) 안에서는 몇 번이든 쓸 수 있고, 지나면
    죽는다. 목적은 «사번이 URL에 영구히 남지 않게»이지 «단 한 번»이 아니었다.
    """
    with _lock:
        emp_id, exp = _TICKETS.get(tok, (None, 0.0))
    return emp_id if emp_id and exp >= time.time() else None


# 다른 라우터(범위별 내보내기)도 같은 표를 쓴다 — 창구가 둘이면 갈라진다.
def check_download_ticket(tok: str | None) -> str | None:
    return _check_ticket(tok) if tok else None


_qmg_path:    Path | None = None
_welder_path: Path | None = None
_qmg_name:    str         = ""
_welder_name: str         = ""
_wps_name:    str         = ""
_consumable_name: str     = ""

def _update_progress(pct: int, msg: str) -> None:
    with _lock:
        _state["progress"] = pct
        _state["message"]  = msg

# ─── 상태 조회 ────────────────────────────────────────────────────────────────
def _load_consumables() -> list[str] | None:
    """승인 용접봉 Lot 목록. 없으면 None — Check 28이 «대상 없음»이 된다."""
    if config.CONSUMABLE_CACHE.exists():
        try:
            d = json.loads(config.CONSUMABLE_CACHE.read_text(encoding='utf-8'))
            return d.get("lots") if isinstance(d, dict) else d
        except Exception:
            logger.exception("승인 용접봉 목록을 읽지 못했다")
    return None


def _load_wps_cache() -> dict | None:
    if config.WPS_CACHE.exists():
        try:
            return json.loads(config.WPS_CACHE.read_text(encoding='utf-8'))
        except Exception:
            return None
    return None

# ─── 상태 영속화 (서버 재시작 후에도 업로드 파일명/검증 결과 유지) ────────────────
#
# 저장소는 `data/qc.db`(SQLite)다. 예전에는 `app_state.json` 한 덩어리였는데,
# 30만 행에서 600MB가 되어 성립하지 않는다(개선계획 1번).
#
# **이름 저장과 결과 저장을 나눈 것이 요점이다.** 예전 `_save_app_state()`는 파일명만
# 바뀌는 업로드에도 결과 20MB 전체를 재직렬화했다 — 30만 건이면 파일 하나 올릴 때마다
# 6초를 버린다. 업로드 경로는 이제 `_save_names()`만 부른다.
def _save_names() -> None:
    """업로드 파일명만 저장한다. 검증 결과는 건드리지 않는다."""
    try:
        result_store.save_names(_qmg_name, _welder_name, _wps_name)
    except Exception:
        logger.exception("업로드 파일명 저장 실패")

def _save_result(result: dict | None) -> None:
    """검증 결과 한 회차를 저장한다. 검증이 끝났을 때만 부른다.

    `_state`에서 읽지 않고 **인자로 받는다.** 저장은 화면에 내보내기 «전에»,
    즉 `_state["result"]`에 넣기 전에 끝나야 하기 때문이다(호출부 주석 참조).
    """
    if result is None:
        return
    result_store.save_run(result)

def _load_app_state() -> None:
    """서버(모듈) 시작 시 이전 업로드 파일/검증 결과를 복원."""
    global _qmg_path, _welder_path, _qmg_name, _welder_name, _wps_name

    qmg_file    = config.UPLOADS_DIR / "qmg_data.xlsx"
    welder_file = config.UPLOADS_DIR / "welder_list.xlsx"
    if qmg_file.exists():
        _qmg_path = qmg_file
    if welder_file.exists():
        _welder_path = welder_file

    # 예전 JSON이 있고 DB가 비어 있으면 1회 옮겨 담는다. JSON은 지우지 않는다.
    try:
        result_store.migrate_from_json()
    except Exception:
        logger.exception("app_state.json 이관 실패 - DB만으로 계속한다")

    try:
        names = result_store.load_names()
        # row_raw는 싣지 않는다 — payload의 64.2%인데 화면은 사용자가 클릭한
        # 한 행만 쓴다. 필요할 때 /api/result/row/{row}로 한 건씩 가져간다.
        result = result_store.load_latest(include_row_raw=False,
                                          include_unworked=False,
                                          include_flags=False)
        saved_at = result_store.latest_saved_at()
    except Exception:
        logger.exception("앱 상태 복원 실패")
        return

    _qmg_name    = names["qmg_name"]
    _welder_name = names["welder_name"]
    _wps_name    = names["wps_name"]
    if result is not None:
        excel_path = config.RESULTS_DIR / "QC_Result.xlsx"
        # 파일이 있다는 것만으로 ready로 복원하면, 검증 N은 성공했지만 Excel 생성이 실패한 뒤
        # 재시작했을 때 검증 N-1의 Excel을 내려주게 된다(다운로드 경로가 경계하던 stale 문제를
        # 재시작이 우회한다). 결과 저장 시각보다 오래된 Excel은 이 결과의 산출물이 아니다.
        excel_ready = False
        if excel_path.exists() and saved_at is not None:
            try:
                excel_ready = excel_path.stat().st_mtime >= saved_at
            except OSError:
                excel_ready = False
            if not excel_ready:
                logger.warning(
                    "결과 Excel이 저장된 검증 결과보다 오래됨 - 재생성 필요 상태로 복원한다"
                )
            elif excel_ready:
                # ZIP(xlsx) 구조 최소 검증: 끝 65KB 안에 EOCD 시그니처가 있어야 한다.
                # 프로세스 강제 종료 시 write_only save()가 중단되면 EOCD 없이 파일이 남는다.
                try:
                    with open(excel_path, 'rb') as _fz:
                        _fz.seek(0, 2)
                        _fz_size = _fz.tell()
                        _search = min(65536, _fz_size)
                        _fz.seek(-_search, 2)
                        _tail = _fz.read(_search)
                    _eocd_sig = b'\x50\x4b\x05\x06'
                    if _eocd_sig not in _tail:
                        logger.warning(
                            "결과 Excel ZIP 구조 손상 (EOCD 없음) — 삭제 후 재생성 필요 상태로 복원"
                        )
                        try:
                            excel_path.unlink()
                        except OSError:
                            pass
                        excel_ready = False
                except OSError:
                    excel_ready = False
        with _lock:
            _state["result"]       = result
            _state["status"]       = "done"
            _state["progress"]     = 100
            _state["message"]      = "검증 완료(이전 결과 복원됨)"
            _state["excel_status"] = "ready" if excel_ready else "pending"
        logger.info("이전 검증 결과 복원 완료 (플래그 %d행)", result.get("flagged_rows", 0))

_load_app_state()

@router.get("/status")
def get_status() -> dict:
    cache = _load_wps_cache()
    consumables = _load_consumables()
    with _lock:
        return {
            "status":          _state["status"],
            "progress":        _state["progress"],
            "message":         _state["message"],
            "error":           _state["error"],
            "has_result":      _state["result"] is not None,
            "excel_status":    _state["excel_status"],
            "excel_error":     _state["excel_error"],
            "has_qmg":         _qmg_path is not None and _qmg_path.exists(),
            "has_welder":      _welder_path is not None and _welder_path.exists(),
            "has_wps_ref":     config.WPS_REF.exists(),
            "has_wps_cache":   cache is not None,
            "wps_cache_count": len(cache) if cache else 0,
            "wps_cache_name":  _wps_name,
            # 승인 용접봉 리스트 — **검증을 막지 않는다.** 없으면 Check 28만
            # «대상 없음»이 된다. 검증 버튼 조건에 넣으면 목록을 아직 못 받은
            # 프로젝트가 아무 검증도 못 하게 된다.
            "has_consumables":   consumables is not None,
            "consumable_count":  len(consumables) if consumables else 0,
            "consumable_name":   _consumable_name,
            "qmg_name":        _qmg_name,
            "welder_name":     _welder_name,
        }

# ─── 요청 가드 ────────────────────────────────────────────────────────────────
def _reject_if_busy() -> None:
    """검증 실행 중 / 결과 Excel 생성 중이면 업로드를 막는다.

    write_result_excel()은 결과가 확정된 **뒤에** qmg_uploads/qmg_data.xlsx를 다시 연다.
    그 사이에 파일이 교체되면, 예전 결과의 절대 행 번호로 새 워크북에 색을 칠하게 된다.
    예외가 나지 않고 조용히 틀린 산출물이 만들어지므로 반드시 입구에서 막아야 한다.
    (Windows CPython의 open()은 공유 모드라 PermissionError가 대신 막아주지 않는다.)
    """
    with _lock:
        if _state["status"] == "running":
            raise HTTPException(409, "검증이 실행 중입니다. 완료 후 다시 업로드하세요")
        if _state["excel_status"] == "generating":
            raise HTTPException(
                409, "이전 검증의 결과 Excel을 생성 중입니다. 잠시 후 다시 시도하세요"
            )


# ─── 파일 업로드 ──────────────────────────────────────────────────────────────
@router.post("/files/qmg")
async def upload_qmg(
    file: UploadFile, x_employee_id: str | None = Header(None)
) -> dict:
    global _qmg_path, _qmg_name
    require_login(x_employee_id)
    _reject_if_busy()
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(400, "xlsx 파일만 허용됩니다")
    config.UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest    = config.UPLOADS_DIR / "qmg_data.xlsx"
    content = await file.read()
    dest.write_bytes(content)
    _qmg_path = dest
    _qmg_name = file.filename
    _save_names()
    logger.info("QMG 업로드: %s (%d bytes)", file.filename, len(content))
    return {"ok": True, "filename": file.filename, "size": len(content)}

@router.post("/files/welder")
async def upload_welder(
    file: UploadFile, x_employee_id: str | None = Header(None)
) -> dict:
    global _welder_path, _welder_name
    require_login(x_employee_id)
    _reject_if_busy()
    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        raise HTTPException(400, "xlsx 파일만 허용됩니다")
    config.UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    dest    = config.UPLOADS_DIR / "welder_list.xlsx"
    content = await file.read()
    dest.write_bytes(content)

    # **올리는 자리에서 바로 읽어 본다.** 검증할 때까지 미루면 «형식이 안 맞는
    # 파일»을 올려놓고도 모른다 — 그 상태로 돌리면 전원이 «자격 미등록»으로
    # 잡히는데, 오류가 아니라 «갑자기 결함이 수천 건»으로만 보인다.
    probe = svc.QCDataLoader()
    try:
        layout = probe.load_welders(dest)
    except Exception as exc:
        dest.unlink(missing_ok=True)
        logger.exception("용접사 파일 파싱 실패: %s", file.filename)
        raise HTTPException(400,
            f"엑셀 파일을 열 수 없습니다 (손상되었거나 지원하지 않는 형식): {exc}")
    if not layout:
        dest.unlink(missing_ok=True)
        raise HTTPException(400,
            "용접사 번호와 Process를 한 건도 찾지 못했습니다 — "
            "새 형식(11행부터 · B열 번호 · D열 Process)이나 "
            "기존 형식(4행부터 · C열 번호 · F열 Process) 중 하나여야 합니다")

    _welder_path = dest
    _welder_name = file.filename
    _save_names()
    n_exp = sum(1 for r in probe.welder_expiry.values() if r["effective"])
    n_blank = len(probe.welder_expiry) - n_exp
    logger.info("Welder 업로드: %s (%d bytes · %s · %d명%s)",
                file.filename, len(content), layout, len(probe.welder_qual),
                f" · 만료일 {n_exp}명 · 누락 {n_blank}명" if probe.welder_expiry else "")
    return {"ok": True, "filename": file.filename, "size": len(content),
            # 어느 형식으로 읽었는지 **화면에 보여준다.** 안 보여주면 옛 파일을
            # 올려놓고 «왜 만료일 검사가 안 도나»를 되짚을 방법이 없다.
            "layout": layout,
            "welders": len(probe.welder_qual),
            "has_expiry": bool(probe.welder_expiry),
            "expiry_dated": n_exp,
            "expiry_missing": n_blank}

@router.post("/files/wps")
async def upload_wps(
    file: UploadFile, x_employee_id: str | None = Header(None)
) -> dict:
    """WPS List xlsx 업로드 → 파싱 후 data/wps_cache.json 에 저장."""
    global _wps_name
    require_login(x_employee_id)
    _reject_if_busy()
    fname = file.filename or ""
    if not fname.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "xlsx 또는 xlsm 파일만 허용됩니다")
    config.UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    tmp = config.UPLOADS_DIR / ("wps_upload" + Path(fname).suffix)
    content = await file.read()
    tmp.write_bytes(content)
    try:
        loader = svc.QCDataLoader()
        try:
            loader.load_wps(tmp)
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("WPS 업로드 파싱 실패: %s", fname)
            raise HTTPException(400, f"엑셀 파일을 열 수 없습니다 (손상되었거나 지원하지 않는 형식): {exc}")
        if not loader.wps_table:
            raise HTTPException(400,
                "'Hull & Topside WPS List' 시트가 없거나 데이터가 없습니다")
        config.WPS_CACHE.parent.mkdir(parents=True, exist_ok=True)
        config.WPS_CACHE.write_text(
            json.dumps(loader.wps_table, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )
        _wps_name = fname
        _save_names()
        count = len(loader.wps_table)
        logger.info("WPS 업로드·캐시 저장: %s (%d개)", fname, count)
        return {"ok": True, "filename": fname, "size": len(content), "wps_count": count}
    finally:
        if tmp.exists():
            tmp.unlink(missing_ok=True)

@router.post("/files/consumable")
async def upload_consumable(
    file: UploadFile, x_employee_id: str | None = Header(None)
) -> dict:
    """승인 용접봉 리스트 업로드 -> Lot No.를 걷어 `consumable_cache.json`에 저장.

    양식은 현업이 준 것에 고정돼 있다 — **25열 · 머리말 3~4행 · 데이터 5행부터 ·
    3열이 Lot No.**(`QCDataLoader.CONSUMABLE_*`).

    **한 건도 못 걷으면 거절한다.** 저장해 두면 화면이 «등록됨»으로 읽는데
    Check 28은 목록이 비어 한 행도 판정하지 않는다 — 「올렸는데 아무것도 안
    잡힌다」가 되고, 원인이 «빈 목록»이라는 것은 알 방법이 없다.
    """
    global _consumable_name
    require_login(x_employee_id)
    _reject_if_busy()
    fname = file.filename or ""
    if not fname.lower().endswith(('.xlsx', '.xlsm')):
        raise HTTPException(400, "xlsx 또는 xlsm 파일만 허용됩니다")
    config.UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
    tmp = config.UPLOADS_DIR / ("consumable_upload" + Path(fname).suffix)
    tmp.write_bytes(await file.read())
    try:
        loader = svc.QCDataLoader()
        try:
            loader.load_consumables(tmp)
        except HTTPException:
            raise
        except Exception as exc:
            logger.exception("승인 용접봉 리스트 파싱 실패: %s", fname)
            raise HTTPException(400,
                f"엑셀 파일을 열 수 없습니다 (손상되었거나 지원하지 않는 형식): {exc}")
        if not loader.consumable_lots:
            raise HTTPException(400,
                f"{svc.QCDataLoader.CONSUMABLE_LOT_COL}번째 열에서 Lot No.를 "
                f"하나도 찾지 못했습니다 — 데이터가 "
                f"{svc.QCDataLoader.CONSUMABLE_HEADER_ROWS + 1}행부터 시작하고 "
                f"{svc.QCDataLoader.CONSUMABLE_LOT_COL}열에 Lot No.가 있는지 "
                f"확인하세요")
        config.CONSUMABLE_CACHE.parent.mkdir(parents=True, exist_ok=True)
        config.CONSUMABLE_CACHE.write_text(
            json.dumps({"filename": fname,
                        "lots": sorted(loader.consumable_lots)},
                       ensure_ascii=False, indent=2),
            encoding='utf-8')
        _consumable_name = fname
    finally:
        tmp.unlink(missing_ok=True)
    logger.info("승인 용접봉 리스트 업로드: %s (Lot %d개)",
                fname, len(loader.consumable_lots))
    return {"ok": True, "filename": fname, "count": len(loader.consumable_lots)}


@router.delete("/files/consumable")
def delete_consumable_cache(x_employee_id: str | None = Header(None)) -> dict:
    """승인 목록을 지운다. **Check 28이 «대상 없음»으로 돌아간다** — 그건
    «승인된 것만 썼다»가 아니라 «대조하지 않았다»다."""
    require_login(x_employee_id)
    _reject_if_busy()
    global _consumable_name
    if config.CONSUMABLE_CACHE.exists():
        config.CONSUMABLE_CACHE.unlink()
    _consumable_name = ""
    logger.info("승인 용접봉 목록 삭제 — Check 28은 다음 검증부터 대상 없음")
    return {"ok": True}


@router.delete("/files/wps")
def delete_wps_cache(x_employee_id: str | None = Header(None)) -> dict:
    """저장된 WPS 캐시 삭제 → 다음 검증부터 xlsm 파일 사용."""
    require_login(x_employee_id)
    _reject_if_busy()
    global _wps_name
    if config.WPS_CACHE.exists():
        config.WPS_CACHE.unlink()
    _wps_name = ""
    _save_names()
    return {"ok": True}

# ─── 검증 실행 ────────────────────────────────────────────────────────────────
@router.post("/verify")
def run_verify(project: str | None = None,
               x_employee_id: str | None = Header(None)) -> dict:
    """검증 실행. `project`는 **어느 프로젝트 기준으로 판정할지**다.

    주지 않으면 지금 활성인 프로파일을 쓴다. 이 값은 결과에 그대로 박히므로
    (`run.project` + `run.rules_json`), 나중에 설정을 바꿔도 «이 결과가 무엇으로
    판정됐는지»가 남는다. 예전에는 이 기록이 없어서 다음이 가능했다:

        A 프로젝트로 검증 -> 세팅에서 B로 전환 -> 화면은 A의 숫자인데
        보는 사람은 B 결과로 읽는다 (오류도 경고도 없다)
    """
    global _qmg_path, _welder_path
    require_login(x_employee_id)
    # **없는 프로젝트 이름은 거절한다.** 여기서 새로 만들면 오타 한 글자가
    # «전 항목 켬»이 되어 엉뚱한 기준으로 돌아가고, 결과에는 그 오타 이름이
    # 박혀서 나중에도 알기 어렵다. 새로 만드는 길은 '프로젝트 세팅' 탭 하나뿐이다.
    if (project or "").strip() and not rule_catalog.has_profile(project):
        known = ", ".join(rule_catalog.profile_names())
        raise HTTPException(400, f"'{project}' 프로젝트가 없습니다. "
                                 f"'프로젝트 세팅'에서 먼저 만드세요 (있는 것: {known})")
    if not _qmg_path or not _qmg_path.exists():
        raise HTTPException(400, "QMG 데이터 파일을 먼저 업로드하세요")
    if not _welder_path or not _welder_path.exists():
        raise HTTPException(400, "Welder List 파일을 먼저 업로드하세요")
    wps_cache = _load_wps_cache()
    if not wps_cache and not config.WPS_REF.exists():
        # WPS 캐시가 업로드되어 있으면 .xlsm 고정 참조 파일이 없어도 검증 가능
        raise HTTPException(500, f"WPS 참조 파일을 찾을 수 없음: {config.WPS_REF}")

    with _lock:
        if _state["status"] == "running":
            raise HTTPException(409, "검증이 이미 실행 중입니다")
        if _state["excel_status"] == "generating":
            raise HTTPException(409, "이전 검증의 Excel 결과 파일을 생성 중입니다. 잠시 후 다시 시도하세요")
        _state.update({
            "status": "running", "progress": 0,
            "message": "시작 중…", "error": None,
            "result": None,
            "excel_status": "idle", "excel_error": None,
        })

    qmg    = _qmg_path
    welder = _welder_path
    qmg_fn    = _qmg_name
    welder_fn = _welder_name
    run_project = (project or "").strip() or None

    def _run() -> None:
        try:
            wps_cache = _load_wps_cache()
            result = svc.run_verification(
                qmg_path=qmg,
                welder_path=welder,
                wps_ref_path=config.WPS_REF,
                qmg_filename=qmg_fn,
                welder_filename=welder_fn,
                wps_cache=wps_cache,
                consumable_lots=_load_consumables(),
                project=run_project,
                progress=_update_progress,
            )
        except Exception as exc:
            logger.exception("검증 오류")
            with _lock:
                _state["status"]  = "error"
                _state["error"]   = str(exc)
                _state["message"] = f"오류: {exc}"
            return

        # ── 저장을 «done» 이전에 끝내고, 화면에는 «저장소가 돌려준 것»을 낸다 ──
        #
        # 화면은 상태를 폴링하다가 done을 보는 즉시 /api/result를 부른다. done을
        # 먼저 세우면 그 틈에 무거운 응답이 나간다 — 재현이 어렵고 «가끔 느리다»로만
        # 드러나는 경쟁이다.
        #
        # **화면에 내보내는 것은 반드시 저장소를 한 번 거친 결과여야 한다.**
        # 판정 엔진이 만든 dict를 그대로 내보내면 «방금 검증한 직후»와 «재시작 후
        # 복원»이 다른 모양이 된다. 실제로 두 번 그렇게 어긋났다(row_raw 64%,
        # 부서/팀 결함 목록 44%). 한쪽만 고치면 다른 쪽이 조용히 남는다.
        #
        # **저장이 실패하면 «완료»라고 말하지 마라.** 예전에는 `_save_result()`가
        # 예외를 삼켰고, 그러면 `load_latest()`가 **직전 회차**를 돌려준다 —
        # `or result` 폴백은 None일 때만 걸리므로 여기 안 걸린다. 화면에는
        # 「검증 완료」와 함께 옛 숫자가 뜨고, 같은 순간 만들어지는 Excel은
        # 새 숫자다. 실측으로 화면 4,508건 · Excel 183건이 동시에 나왔다.
        # 오류도 경고도 없어서 «아까 그 숫자가 아닌데»로만 보인다.
        try:
            _save_result(result)
            published = result_store.load_latest(include_row_raw=False,
                                                 include_unworked=False,
                                                 include_flags=False)
            if published is None:
                raise RuntimeError("저장 직후 회차를 다시 읽지 못했습니다")
            # 방금 저장한 그 회차인가. 아니면 옛 회차를 공표하려는 참이다.
            if published.get("run_at") != result.get("run_at"):
                raise RuntimeError(
                    "저장된 회차가 방금 판정한 것과 다릅니다 "
                    f"(저장소 {published.get('run_at')!r} != 판정 {result.get('run_at')!r})")
        except Exception as exc:
            logger.exception("검증 결과 저장 실패")
            with _lock:
                _state["status"]  = "error"
                _state["error"]   = str(exc)
                _state["message"] = f"결과를 저장하지 못했습니다: {exc}"
            return

        # 검사 결과는 즉시 확정 → Excel 생성은 사용자가 버튼을 눌렀을 때 시작
        with _lock:
            _state["result"]       = published
            _state["status"]       = "done"
            _state["progress"]     = 100
            _state["message"]      = "검증 완료"
            _state["excel_status"] = "pending"
            _state["excel_error"]  = None

        # 이전 검증으로 만들어둔 범위별(부서/팀) 파일은 전부 낡은 것이 된다.
        try:
            from backend.routers import exports as _exports
            _exports.invalidate()
        except Exception:
            logger.exception("범위별 내보내기 캐시 비우기 실패")

    threading.Thread(target=_run, daemon=True).start()
    return {"status": "running", "message": "검증을 시작했습니다"}

# ─── Excel 생성 (on-demand) ──────────────────────────────────────────────────
@router.post("/excel/build")
def start_excel_build(x_employee_id: str | None = Header(None)) -> JSONResponse:
    """검증 완료 후 사용자가 원할 때 Excel 결과 파일 생성을 시작한다."""
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(400, "검증이 완료되지 않았습니다")
        current = _state["excel_status"]
        if current in ("generating", "ready"):
            return JSONResponse({"excel_status": current})
        _state["excel_status"] = "generating"
        _state["excel_error"]  = None
        qmg      = _qmg_path
        res      = _state["result"]
        feedback = feedback_store.load_feedback()

    def _gen_excel() -> None:
        try:
            config.RESULTS_DIR.mkdir(parents=True, exist_ok=True)
            excel_report.build_report(
                qmg, res, config.RESULTS_DIR / "QC_Result.xlsx",
                feedback=feedback,
                priority=svc._PRIORITY,
            )
            with _lock:
                _state["excel_status"] = "ready"
        except Exception as exc:
            logger.exception("Excel 결과 생성 오류")
            with _lock:
                _state["excel_status"] = "error"
                _state["excel_error"]  = str(exc)

    threading.Thread(target=_gen_excel, daemon=True).start()
    return JSONResponse({"excel_status": "generating"})

# ─── 결과 조회 ────────────────────────────────────────────────────────────────
@router.get("/result")
def get_result(x_employee_id: str | None = Header(None)) -> JSONResponse:
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
        return JSONResponse(_state["result"])

def _id_list(raw: str | None) -> list[str] | None:
    """`"4,17,RUYA1"` 같은 질의 문자열을 **식별자 목록**으로.

    None(파라미터 미지정)과 `[]`(전부 해제)를 구분해야 한다 — 후자는 결과 0건이
    맞는 답이고, 이걸 «필터 없음»으로 뭉개면 화면에서 Check를 전부 해제했는데
    전체가 나오는 상태가 된다.

    **정수로 바꾸지 마라.** 두 가지가 조용히 깨진다:

      · `RUYA1`이 `isdigit()`에서 걸러져 **프로젝트 규칙을 영원히 못 고른다**
      · DB의 `check_no`가 TEXT라 정수 파라미터가 한 행도 안 맞아 **전부 0건**이 된다

    `-1`은 «전부 해제» 신호라 그대로 통과시킨다.
    """
    if raw is None:
        return None
    return [x.strip() for x in raw.split(",") if x.strip()]


def _int_list(raw: str | None) -> list[int] | None:
    """행 번호처럼 **정말로 정수인** 목록. Check 식별자에는 쓰지 마라."""
    if raw is None:
        return None
    return [int(x) for x in raw.split(",") if x.strip().lstrip("-").isdigit()]


@router.get("/flags")
def get_flags(check: str | None = None, checks: str | None = None,
              dept: str | None = None, team: str | None = None,
              q: str = "", rows: str | None = None, welder: str | None = None,
              drawing_no: str | None = None, joint_no: str | None = None,
              offset: int = 0, limit: int = 50,
              x_employee_id: str | None = Header(None)) -> dict:
    """플래그 목록 — 조건에 맞는 한 페이지.

    예전에는 `/api/result`가 전량을 실어 보냈다(실측 2,521KB, payload의 84%).
    30만 행이면 ~76MB라 브라우저가 받지 못한다. 한 화면이 쓰는 것은 한 페이지뿐이다.

    `limit=0`이면 개수만 센다(items 없음) — 화면이 «몇 건인지»만 필요할 때 쓴다.
    """
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
    return result_store.query_flags(
        check=check, checks=_id_list(checks), dept=dept, team=team, q=q,
        rows=_int_list(rows), welder=welder,
        drawing_no=drawing_no, joint_no=joint_no,
        offset=max(0, offset), limit=max(0, min(limit, config.MAX_FLAG_PAGE)),
    )


@router.get("/flags/aggregate")
def get_flags_aggregate(checks: str | None = None, dept: str | None = None,
                        team: str | None = None,
                        x_employee_id: str | None = Header(None)) -> dict:
    """선택된 Check만으로 다시 낸 대시보드 수치.

    화면의 «Check 취사선택»이 브라우저에서 전체 배열을 훑어 하던 계산이다.
    30만 행이면 그 배열 자체가 오지 못하므로 서버에서 센다.
    """
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
    return result_store.aggregate_flags(
        checks=_id_list(checks), dept=dept, team=team)


@router.get("/flags/facets")
def get_flag_facets(dept: str | None = None,
                    x_employee_id: str | None = Header(None)) -> dict:
    """필터 select 값 목록. 화면이 전체 flags를 훑어 Set으로 모으던 것이다.

    정렬은 한국어 기준이라 SQL로 재현할 수 없어 **정렬하지 않고 보낸다** —
    화면이 지금처럼 `localeCompare(_, 'ko')`로 정렬한다.
    """
    require_login(x_employee_id)
    return result_store.flag_facets(dept=dept)


@router.get("/flags/pattern-teams")
def get_pattern_teams(x_employee_id: str | None = Header(None)) -> dict:
    """(용접사, Check, 부서) -> 팀 목록. 반복 패턴 표의 '팀' 열이 쓴다.

    화면이 표를 그리는 도중에 동기로 훑던 값이라 패턴마다 요청할 수 없다.
    조합 전체를 한 번에 준다.
    """
    require_login(x_employee_id)
    return {"items": result_store.pattern_teams()}


@router.get("/flags/review-counts")
def get_review_counts(welder: str | None = None, by: str | None = None,
                      x_employee_id: str | None = Header(None)) -> dict:
    """검토 기록이 없는 플래그 건수.

    `by=welder`면 **용접사별로 한 번에** 낸다 — 용접사 목록의 '미검토' 열이
    146명분을 동시에 쓰므로 한 명씩 물으면 146번을 왕복하게 된다.

    임시부재(1C)는 판정 대상이 아니라 플래그가 없고, 기록 파일에 남은 옛
    1C 검토 기록도 `review_counts()`가 항상 거른다 — 선택이 아니다.
    """
    require_login(x_employee_id)
    fb = feedback_store.load_feedback()
    if by == "welder":
        return {"by_welder": result_store.review_counts_by_welder(fb)}
    return result_store.review_counts(fb, welder=welder)


@router.get("/departments/defects")
def get_dept_defects(dept: str, team: str | None = None, q: str = "",
                     offset: int = 0, limit: int = 50,
                     x_employee_id: str | None = Header(None)) -> dict:
    """부서/팀 결함 목록 — 한 페이지씩.

    4단계에서는 화면이 `flags`로 직접 만들었는데(전체 배열이 있다는 전제),
    그 전제가 없어지므로 서버가 낸다. `expand_defects()`와 같은 결과다.

    **부서명을 경로가 아니라 질의 파라미터로 받는다.** 실데이터에 슬래시가 든
    부서명이 있어(`Q530-SMR/ITER생산부`) 경로에 넣으면 라우트가 쪼개져 404가 된다.
    집계는 질의 파라미터라 멀쩡한데 목록만 빈 표가 나오는, 찾기 어려운 형태였다.
    """
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
    return result_store.dept_defects(dept, team=team, q=q,
                                     offset=max(0, offset),
                                     limit=max(1, min(limit, 2000)))


@router.get("/result/unworked")
def get_result_unworked(offset: int = 0, limit: int = 50, q: str = "",
                        x_employee_id: str | None = Header(None)) -> dict:
    """미작업(Check 0) 목록 — 한 페이지씩.

    예전에는 `/api/result`가 전체를 실어 보냈다(실측 590KB, payload의 17%).
    화면은 모달을 열었을 때만, 그것도 한 페이지만 쓴다. 30만 행이면 ~18MB다.
    """
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
    limit = max(1, min(limit, 500))
    return result_store.get_unworked(offset=max(0, offset), limit=limit, q=q)


@router.get("/result/row/{row_no}")
def get_result_row(row_no: int,
                   x_employee_id: str | None = Header(None)) -> dict:
    """한 행의 엑셀 원문 스냅샷 (상세 모달 '원본 데이터' 패널).

    예전에는 `/api/result`가 전 행의 스냅샷을 통째로 실어 보냈다(payload의 64.2%).
    화면은 사용자가 클릭한 한 행만 쓰므로 그때 한 건씩 가져간다.
    """
    require_login(x_employee_id)
    with _lock:
        if _state["status"] != "done" or _state["result"] is None:
            raise HTTPException(404, "결과 없음. 먼저 검증을 실행하세요")
    raw = result_store.get_row_raw(row_no)
    if raw is None:
        raise HTTPException(404, f"{row_no}행의 원본 스냅샷이 없습니다 "
                                 f"(결함/미작업 행이 아니거나 이전 검증의 행 번호)")
    # 저장에는 **값만** 들어 있다(머리말은 행마다 같아서 뺐다 — payload의 62%였다).
    # 화면이 기대하는 `{label, col, value}` 모양은 여기서 다시 만든다. 사용자가
    # 클릭한 한 행뿐이라 비용이 없다. **빠뜨리면 상세 모달이 통째로
    # `undefined, undefined열`이 된다** — 오류는 나지 않는다.
    return {"row": row_no, "raw": svc.raw_with_labels(raw)}


def get_qmg_path() -> Path | None:
    """다른 라우터(범위별 내보내기)가 원본 파일 경로를 참조하기 위한 접근자."""
    return _qmg_path


def get_last_result() -> dict | None:
    """다른 라우터(설정/부서 발송 등)에서 최근 검증 결과를 참조하기 위한 접근자."""
    with _lock:
        return _state["result"]

class WelderXlsxIn(BaseModel):
    welder_nos: list[str]
    label: str = "전체"


@router.post("/welders/excel")
def welders_excel(body: WelderXlsxIn,
                  x_employee_id: str | None = Header(None)) -> Response:
    """용접사 목록을 Excel로 낸다 — **화면이 보내 준 번호만** 채운다.

    서버가 다시 거르지 않는 것이 요점이다. 필터·검색을 여기서 재현하면 화면과
    갈라져 「화면은 71명인데 파일은 146명」이 된다 — 받는 사람은 어느 쪽이 맞는지
    알 수 없다. 화면이 «지금 보고 있는 목록»을 그대로 보내고, 여기서는 그 번호에
    해당하는 값만 붙인다.

    업체에 그대로 보낼 수 있어야 해서 **부서·팀**이 함께 나간다. 번호만으로는
    누구에게 물어야 할지 알 수 없다. 한 사람이 여러 팀에 걸치면 **줄을 나눈다** —
    한 칸에 몰아넣으면 팀별로 거를 수가 없다.
    """
    require_login(x_employee_id)
    result = get_last_result()
    if result is None:
        raise HTTPException(404, "검증 결과가 없습니다. 먼저 검증을 실행하세요")
    info = result.get("welder_info") or {}
    wanted = [w for w in body.welder_nos if w in info]
    if not wanted:
        raise HTTPException(400, "내보낼 용접사가 없습니다")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "용접사"
    head = ["용접사 번호", "이름", "부서", "팀", "그 팀에서 시공", "자격 등록",
            "보유 자격", "시공 행", "결함 행", "결함률(%)", "미자격 지적(C6)"]
    ws.append(head)
    for no in wanted:
        w = info[no]
        orgs = w.get("orgs") or [{"department": "", "team": "", "rows": 0}]
        for o in orgs:
            ws.append([
                no, w.get("name", ""), o.get("department", ""), o.get("team", ""),
                o.get("rows", 0),
                "등록" if w.get("registered") else "미등록",
                ", ".join(w.get("processes") or []),
                w.get("total", 0), w.get("defect", 0), w.get("rate", 0),
                w.get("missing_flags", 0),
            ])

    fill = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
        # 열 너비는 실제 내용에 맞춘다 — 잘려 보이면 받는 사람이 되묻는다
        width = max((len(str(ws.cell(row=r, column=c).value or ""))
                     for r in range(1, min(ws.max_row, 500) + 1)), default=8)
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 4, 10), 40)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    name = f"용접사_{body.label}_{result.get('run_at', '')[:10]}.xlsx"
    logger.info("용접사 목록 Excel: %s (%d명 / %d줄)",
                body.label, len(wanted), ws.max_row - 1)
    return Response(
        content=data, media_type=_XLSX,
        headers={
            "Content-Disposition":
                f"attachment; filename*=UTF-8''{urllib.parse.quote(name)}",
            "Cache-Control": "no-store, no-cache, must-revalidate",
        })


@router.post("/result/excel/ticket")
def issue_excel_ticket(x_employee_id: str | None = Header(None)) -> dict:
    """다운로드용 표를 발급한다. 30초 동안 유효하며 그 안에서는 재사용된다."""
    return {"ticket": _issue_ticket(require_login(x_employee_id))}


@router.get("/result/excel")
def download_excel(
    t: str | None = None,
    x_employee_id: str | None = Header(None),
) -> FileResponse:
    # 헤더가 있으면 그걸로, 없으면 «일회용 표»로 확인한다.
    # 표는 `<a download>` 경로 전용이다 — 그쪽은 헤더를 실을 수 없다.
    if x_employee_id:
        require_login(x_employee_id)
    elif not (t and _check_ticket(t)):
        raise HTTPException(401, "다운로드 표가 없거나 만료되었습니다. 다시 시도하세요")
    excel_path = config.RESULTS_DIR / "QC_Result.xlsx"
    with _lock:
        status       = _state["status"]
        excel_status = _state["excel_status"]
    # 새 검증이 진행 중이면(status=running) 디스크에 이전 결과 파일이 남아있어도
    # 그건 지금 표시 중인 결과와 무관한 stale 파일이므로 내려주지 않는다.
    if status == "running":
        raise HTTPException(409, "검증이 진행 중입니다. 완료 후 다시 시도하세요")
    if not excel_path.exists() or excel_status == "generating":
        if excel_status == "generating":
            raise HTTPException(409, "Excel 결과 파일을 생성 중입니다. 잠시 후 다시 시도하세요")
        raise HTTPException(404, "Excel 결과 파일 없음. 먼저 검증을 실행하세요")
    return stream_xlsx(excel_path, "QC_Result.xlsx")
